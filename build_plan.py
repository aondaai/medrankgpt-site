# -*- coding: utf-8 -*-
import unicodedata, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def slug(s):
    s = unicodedata.normalize('NFKD', s).encode('ascii','ignore').decode('ascii')
    s = s.lower()
    s = re.sub(r'[^a-z0-9]+','-',s).strip('-')
    return s

# ---------- paleta ----------
GRAPHITE="14181F"; GOLD="C9A24B"; GOLD_L="F3E9CE"; INK="1B2230"
LIGHT="F6F3EC"; BAND="FAF7F0"; WHITE="FFFFFF"
thin=Side(style="thin", color="DDDDDD")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
F="Arial"
def font(**k): k.setdefault("name",F); return Font(**k)

wb=Workbook()

# ============================================================
# DADOS — 100 PÁGINAS
# ============================================================
rows=[]  # cada item: dict
def add(cluster,tipo,titulo,prompt,kw,intent,esp,regiao,funil,formato,schema,cta,onda,fonte):
    rows.append(dict(cluster=cluster,tipo=tipo,titulo=titulo,slug="/"+slug(titulo),
        prompt=prompt,kw=kw,intent=intent,esp=esp,regiao=regiao,funil=funil,
        formato=formato,schema=schema,cta=cta,onda=onda,fonte=fonte,status="A fazer"))

# Cluster 1 — Pilares / educação (10)
C1="1. Pilares & Educação"
pilares=[
 ("O que é AEO (Answer Engine Optimization) na medicina","o que é AEO para médicos","AEO médico","Informacional","Glossário/Definição","Article + FAQPage + DefinedTerm","Profound: AEO Ultimate Guide"),
 ("O que é GEO (Generative Engine Optimization) para médicos","o que é GEO em saúde","GEO médico","Informacional","Glossário/Definição","Article + FAQPage + DefinedTerm","Profound: AEO Ultimate Guide"),
 ("SEO vs AEO vs GEO: o guia definitivo para médicos","diferença entre SEO AEO e GEO","SEO AEO GEO médico","Informacional","Guia/Comparativo","Article + FAQPage","Noble: SEO isn't dead"),
 ("Como o ChatGPT escolhe qual médico recomendar","como o ChatGPT recomenda médicos","ChatGPT recomenda médico","Informacional","Guia prático (pillar)","Article + FAQPage","Profound: How to rank in ChatGPT"),
 ("Como o Google AI Overviews mostra médicos nas buscas","como aparecer no Google AI Overviews médico","Google AI Overviews médico","Informacional","Guia prático","Article + FAQPage","Profound: Technical AEO"),
 ("Como aparecer no Gemini sendo médico ou clínica","como aparecer no Gemini médico","Gemini para médicos","Informacional","Guia prático","Article + FAQPage","Profound: How to rank in ChatGPT"),
 ("Como ser citado pelo Perplexity: guia para médicos","como ser citado no Perplexity médico","Perplexity médico","Informacional","Guia prático","Article + FAQPage","Noble: 7-step playbook"),
 ("Marketing médico na era da IA: o guia completo","marketing médico inteligência artificial","marketing médico IA","Informacional","Guia pillar","Article + FAQPage","Profound: AEO Ultimate Guide"),
 ("Glossário de IA e busca para médicos (AEO, GEO, LLM, prompt)","glossário IA busca médicos","glossário AEO GEO","Informacional","Glossário","Article + DefinedTermSet","Profound: Research Hub"),
 ("O que é E-E-A-T médico e por que a IA valoriza","o que é EEAT médico","E-E-A-T médico","Informacional","Definição/Guia","Article + FAQPage","Profound: Technical AEO"),
]
for t,p,k,i,f,s,src in pilares:
    add(C1,"Pilar / Definição",t,p,k,i,"Geral","Brasil","Topo (ToFu)",f,s,"Diagnóstico grátis","Onda 1",src)

# Cluster 2 — Como aparecer no ChatGPT por especialidade (16)
C2="2. Especialidade × IA"
esps=["Dermatologia","Cirurgia Plástica","Oftalmologia","Ortopedia","Ginecologia","Cardiologia",
 "Psiquiatria","Pediatria","Endocrinologia","Urologia","Otorrinolaringologia","Gastroenterologia",
 "Nutrologia","Reprodução Humana","Odontologia (HOF)","Medicina Esportiva"]
for e in esps:
    add(C2,"Especialidade × IA",
        f"Como {e.lower()} aparece no ChatGPT e no Google AI",
        f"melhor médico de {e.lower()} segundo a IA",
        f"{e.lower()} ChatGPT","Comercial-informacional",e,"Brasil","Meio (MoFu)",
        "Guia + FAQ por especialidade","Service + FAQPage + Physician","Agendar reunião",
        "Onda 1" if e in ("Dermatologia","Cirurgia Plástica","Oftalmologia","Ortopedia") else "Onda 2",
        "Profound: How to rank in ChatGPT")

# Cluster 3 — Programático local: especialidade × cidade (30)
C3="3. Local (Especialidade × Cidade)"
cidades=["São Paulo","Rio de Janeiro","Belo Horizonte","Brasília","Curitiba","Porto Alegre",
 "Salvador","Recife","Fortaleza","Goiânia"]
local_esp=["Dermatologista","Cirurgião plástico","Ortopedista","Oftalmologista","Ginecologista","Psiquiatra"]
combos=[]
idx=0
for c in cidades:
    for j in range(3):
        e=local_esp[idx%len(local_esp)]; idx+=1
        combos.append((e,c))
for e,c in combos:
    add(C3,"Local programática",
        f"Melhor {e.lower()} em {c}: como a IA escolhe (e como aparecer)",
        f"melhor {e.lower()} em {c}",
        f"{e.lower()} {c}","Comercial (local)",e.replace('ç','c'),c,"Meio/Fundo (MoFu/BoFu)",
        "Listicle local + FAQ","Service + FAQPage + LocalBusiness","Agendar reunião",
        "Onda 2" if cidades.index(c)<5 else "Onda 3","Noble: Top 10 / listicle local")

# Cluster 4 — Condições / sintomas (14)
C4="4. Condições & Sintomas"
cond=[("Acne","Dermatologia"),("Hérnia de disco","Ortopedia"),("Catarata","Oftalmologia"),
 ("Ansiedade","Psiquiatria"),("Refluxo gastroesofágico","Gastroenterologia"),("Varizes","Cirurgia Vascular"),
 ("Enxaqueca","Neurologia"),("Hipotireoidismo","Endocrinologia"),("Infertilidade","Reprodução Humana"),
 ("Rinite alérgica","Otorrinolaringologia"),("Apneia do sono","Otorrinolaringologia"),
 ("Menopausa","Ginecologia"),("Próstata aumentada","Urologia"),("Dor no joelho","Ortopedia")]
for c,e in cond:
    add(C4,"Condição / Sintoma",
        f"{c}: o que a IA responde — e por que consultar um especialista",
        f"{c.lower()} o que fazer",
        f"{c.lower()}","Informacional (paciente)",e,"Brasil","Topo/Meio (ToFu/MoFu)",
        "Resposta direta + FAQ","MedicalWebPage + FAQPage","Diagnóstico grátis","Onda 2",
        "Noble: data-led patient hook")

# Cluster 5 — Procedimentos (8)
C5="5. Procedimentos Eletivos"
proc=[("Rinoplastia","Cirurgia Plástica"),("Cirurgia refrativa (miopia)","Oftalmologia"),
 ("Preenchimento facial","Dermatologia"),("Cirurgia bariátrica","Cirurgia Geral"),
 ("Implante capilar","Dermatologia"),("Toxina botulínica (botox)","Dermatologia"),
 ("Artroscopia de joelho","Ortopedia"),("Lentes de contato dental","Odontologia (HOF)")]
for p,e in proc:
    add(C5,"Procedimento",
        f"{p}: como a IA explica e como ser o médico citado",
        f"{p.lower()} vale a pena",
        f"{p.lower()}","Comercial-informacional",e,"Brasil","Meio (MoFu)",
        "Guia procedimento + FAQ","MedicalProcedure + FAQPage","Agendar reunião","Onda 2",
        "Profound: How to rank in ChatGPT")

# Cluster 6 — Comparativos X vs Y (6)
C6="6. Comparativos (X vs Y)"
comp=[
 "MedRankGPT vs agência de marketing médico tradicional",
 "SEO vs AEO para médicos: qual priorizar primeiro",
 "ChatGPT vs Google: onde o paciente busca médico hoje",
 "Site próprio vs Doctoralia e portais: o que a IA realmente cita",
 "Tráfego pago vs ser citado pela IA: o que rende mais para médicos",
 "Agência vs ferramenta de IA: quem faz o trabalho por você",
]
for t in comp:
    add(C6,"Comparativo",t,t.lower(),"comparativo "+slug(t).replace('-',' ')[:30],
        "Comercial","Geral","Brasil","Meio/Fundo (MoFu/BoFu)","Comparativo + tabela + FAQ",
        "Article + FAQPage","Agendar reunião","Onda 1","Profound: Profound vs X")

# Cluster 7 — Estudos de dados / autoridade (4)
C7="7. Dados & Autoridade (PR)"
dados=[
 ("Índice MedRankGPT: as especialidades mais citadas pela IA no Brasil","índice IA médicos brasil","Estudo proprietário","Dataset + Article","Profound: Profound Index"),
 ("25 estatísticas de IA na saúde que todo médico deveria saber","estatísticas IA saúde","Roundup de dados","Article + FAQPage","Noble: AI Search Statistics 2026"),
 ("Quantas menções a IA precisa para recomendar um médico (estudo)","menções IA recomendação médico","Estudo/análise","Article + Dataset","Noble: how many mentions"),
 ("Onsite vs offsite: de onde vem o que a IA diz sobre médicos","onsite offsite IA médicos","Estudo/tese","Article + FAQPage","Noble: 77% offsite"),
]
for t,k,f,s,src in dados:
    add(C7,"Estudo de dados",t,k,k,"Autoridade/PR","Geral","Brasil","Topo (ToFu)",f,s,
        "Diagnóstico grátis","Onda 1",src)

# Cluster 8 — Conformidade CFM (4)
C8="8. Conformidade CFM"
cfm=[
 "É permitido pelo CFM aparecer na IA? O que pode e o que não pode",
 "Publicidade médica e inteligência artificial: guia de conformidade",
 "O que o Código de Ética Médica diz sobre marketing digital",
 "Erros de marketing médico que ferem o CFM (e como evitar)",
]
for t in cfm:
    add(C8,"Conformidade",t,t.lower()[:45],"CFM marketing médico","Informacional/Confiança",
        "Geral","Brasil","Meio (MoFu)","Guia + FAQ","Article + FAQPage","Falar com especialista",
        "Onda 1" if "permitido" in t else "Onda 2","Original (whitespace)")

# Cluster 9 — Listicles / ferramentas (4)
C9="9. Listicles & Checklists"
lst=[
 ("10 ferramentas de IA para médicos e clínicas em 2026","ferramentas IA médicos","Listicle","Article + ItemList + FAQPage","Noble: Top 10 tools"),
 ("As 7 melhores práticas de AEO para consultórios","melhores práticas AEO consultório","Listicle/Playbook","Article + HowTo + FAQPage","Noble: 7-step playbook"),
 ("Checklist: seu consultório está pronto para a busca por IA?","checklist consultório IA","Checklist","Article + HowTo","Profound: Technical AEO"),
 ("9 erros que deixam médicos invisíveis para a IA","erros médicos invisíveis IA","Listicle/erros","Article + FAQPage","Noble: 7 mistakes"),
]
for t,k,f,s,src in lst:
    add(C9,"Listicle / Checklist",t,k,k,"Informacional","Geral","Brasil","Topo/Meio (ToFu/MoFu)",f,s,
        "Diagnóstico grátis","Onda 2",src)

# Cluster 10 — Casos / prova (2)
C10="10. Casos & Prova"
casos=[
 "Caso: como uma clínica de dermatologia virou citação no ChatGPT",
 "Caso: cirurgião plástico triplicou suas menções em IA em 90 dias",
]
for t in casos:
    add(C10,"Caso de sucesso",t,t.lower()[:45],"caso IA médico","Prova/Decisão",
        "Geral","Brasil","Fundo (BoFu)","Estudo de caso (% lift)","Article (CaseStudy)",
        "Agendar reunião","Onda 2","Noble: case studies")

# Cluster 11 — Processo / transparência (2)
C11="11. Processo & Transparência"
proceso=[
 "Como funciona o MedRankGPT: do diagnóstico à citação na IA",
 "O que acontece depois que você contrata o MedRankGPT",
]
for t in proceso:
    add(C11,"Processo",t,t.lower()[:45],"como funciona MedRankGPT","Decisão",
        "Geral","Brasil","Fundo (BoFu)","Passo a passo + FAQ","Article + HowTo + FAQPage",
        "Agendar reunião","Onda 1","Noble: What happens after you sign up")

assert len(rows)==100, f"esperado 100, obtido {len(rows)}"

# ============================================================
# ABA 2 — PLANO 100 PÁGINAS
# ============================================================
ws=wb.active; ws.title="Plano 100 Páginas"
headers=["ID","Cluster","Tipo / Arquétipo","Título da página (H1)","Slug (URL)",
 "Prompt-alvo (pergunta à IA)","Palavra-chave SEO","Intenção","Especialidade","Região",
 "Funil","Formato AEO","Schema recomendado","CTA","Onda","Inspiração (fonte)","Status"]
# título da aba
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
tc=ws.cell(1,1,"MedRankGPT — Estratégia de Conteúdo: 100 Páginas para Dominar a Busca por IA")
tc.font=font(bold=True,size=14,color="FFFFFF"); tc.fill=PatternFill("solid",fgColor=GRAPHITE)
tc.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=30
# header
hr=2
for j,h in enumerate(headers,1):
    c=ws.cell(hr,j,h); c.font=font(bold=True,color="FFFFFF",size=10)
    c.fill=PatternFill("solid",fgColor=GOLD); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=border
ws.row_dimensions[hr].height=34
# data
for r,d in enumerate(rows,start=hr+1):
    vals=[r-hr,d["cluster"],d["tipo"],d["titulo"],d["slug"],d["prompt"],d["kw"],d["intent"],
          d["esp"],d["regiao"],d["funil"],d["formato"],d["schema"],d["cta"],d["onda"],d["fonte"],d["status"]]
    for j,v in enumerate(vals,1):
        c=ws.cell(r,j,v); c.font=font(size=9.5); c.border=border
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal="center" if j in(1,8,10,11,15,17) else "left")
    if (r-hr)%2==0:
        for j in range(1,len(headers)+1): ws.cell(r,j).fill=PatternFill("solid",fgColor=BAND)
widths=[5,26,20,40,30,34,22,20,18,14,18,24,28,20,9,26,11]
for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
ws.freeze_panes="A3"
ws.auto_filter.ref=f"A2:{get_column_letter(len(headers))}{hr+len(rows)}"

# ============================================================
# ABA 1 — ESTRATÉGIA (inserir como primeira)
# ============================================================
es=wb.create_sheet("Estratégia",0)
es.sheet_view.showGridLines=False
def setw(sh,m):
    for col,w in m.items(): sh.column_dimensions[col].width=w
setw(es,{"A":3,"B":40,"C":16,"D":16,"E":52,"F":16})
def title(sh,row,text,span=6,size=15):
    sh.merge_cells(start_row=row,start_column=2,end_row=row,end_column=1+span)
    c=sh.cell(row,2,text); c.font=font(bold=True,size=size,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=GRAPHITE); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    sh.row_dimensions[row].height=30
def sub(sh,row,text):
    c=sh.cell(row,2,text); c.font=font(bold=True,size=12,color=INK); sh.row_dimensions[row].height=22
def para(sh,row,text,span=5):
    sh.merge_cells(start_row=row,start_column=2,end_row=row,end_column=1+span)
    c=sh.cell(row,2,text); c.font=font(size=10,color="333333"); c.alignment=Alignment(wrap_text=True,vertical="top")

title(es,2,"MedRankGPT · Estratégia de Conteúdo AEO/GEO — 100 Páginas")
es.row_dimensions[3].height=6
para(es,4,"Plano para tornar o MedRankGPT a referência que a IA cita quando médicos e pacientes "
     "perguntam sobre saúde, especialistas e como aparecer na IA. Modelo inspirado nas máquinas de "
     "conteúdo da Profound (unicórnio de AEO) e da Noble (menções offsite), adaptado ao mercado médico brasileiro "
     "e às normas do CFM. Estratégia: 'comer da própria comida' — ser citado pela IA praticando o que vendemos.",span=5)
es.row_dimensions[4].height=58
es.row_dimensions[5].height=42
para(es,5,"Princípio AEO: cada página responde UMA pergunta real (prompt) de forma direta e citável, "
     "com FAQ + dados estruturados (schema). É isso que os modelos de IA extraem e citam.",span=5)

# KPIs
sub(es,7,"Metas (North Star & funil)")
kpis=[("North Star","Share of voice em IA","Nº de prompts/cidades em que a marca/cliente é citado pela IA"),
 ("Tráfego","Visitas orgânicas + de IA","Crescimento mês a mês vindo de Google e de bots de IA"),
 ("Conversão","Diagnósticos solicitados","Leads do CTA 'Diagnóstico grátis' por página"),
 ("Autoridade","Menções/citações","Páginas citadas por ChatGPT, Gemini, Perplexity e AI Overviews")]
hr=8
for j,h in enumerate(["Camada","KPI","Como medir"],0):
    c=es.cell(hr,2+j,h); c.font=font(bold=True,color="FFFFFF",size=10); c.fill=PatternFill("solid",fgColor=GOLD)
    c.alignment=Alignment(horizontal="left",vertical="center"); c.border=border
es.merge_cells(start_row=hr,start_column=4,end_row=hr,end_column=6)
for i,(a,b,cc) in enumerate(kpis):
    rr=hr+1+i
    es.cell(rr,2,a).font=font(size=10,bold=True); es.cell(rr,3,b).font=font(size=10)
    es.merge_cells(start_row=rr,start_column=4,end_row=rr,end_column=6)
    cm=es.cell(rr,4,cc); cm.font=font(size=10); cm.alignment=Alignment(wrap_text=True,vertical="top")
    for cidx in (2,3,4): es.cell(rr,cidx).border=border

# Distribuição por cluster (com COUNTIF dinâmico)
sub(es,14,"Distribuição por cluster (100 páginas)")
clusters=[C1,C2,C3,C4,C5,C6,C7,C8,C9,C10,C11]
desc={C1:"Definições e guias-pilar (topo de funil, autoridade)",
 C2:"Como cada especialidade aparece na IA (alta intenção)",
 C3:"Programáticas locais: especialidade × cidade (escala + busca local)",
 C4:"Condições e sintomas que o paciente pergunta à IA",
 C5:"Procedimentos eletivos de alto valor",
 C6:"Comparativos 'X vs Y' (decisão)",
 C7:"Estudos de dados proprietários (PR + links + citação)",
 C8:"Conformidade com o CFM (confiança, diferencial)",
 C9:"Listicles e checklists (formato que a IA adora citar)",
 C10:"Casos de sucesso com % de ganho",
 C11:"Processo e transparência (fundo de funil)"}
hr=15
for j,h in enumerate(["Cluster","Qtde","Papel estratégico"],0):
    c=es.cell(hr,2+j,h); c.font=font(bold=True,color="FFFFFF",size=10); c.fill=PatternFill("solid",fgColor=GOLD)
    c.alignment=Alignment(horizontal="left",vertical="center"); c.border=border
es.merge_cells(start_row=hr,start_column=4,end_row=hr,end_column=6)
PLAN="'Plano 100 Páginas'"
for i,cl in enumerate(clusters):
    rr=hr+1+i
    es.cell(rr,2,cl).font=font(size=10)
    es.cell(rr,3,f"=COUNTIF({PLAN}!$B$3:$B$102,$B{rr})").font=font(size=10,bold=True)
    es.cell(rr,3).alignment=Alignment(horizontal="center")
    es.merge_cells(start_row=rr,start_column=4,end_row=rr,end_column=6)
    dd=es.cell(rr,4,desc[cl]); dd.font=font(size=10); dd.alignment=Alignment(wrap_text=True,vertical="top")
    for cidx in (2,3,4): es.cell(rr,cidx).border=border
totr=hr+1+len(clusters)
es.cell(totr,2,"TOTAL").font=font(bold=True,size=10,color="FFFFFF"); es.cell(totr,2).fill=PatternFill("solid",fgColor=INK)
tc=es.cell(totr,3,f"=SUM(C{hr+1}:C{hr+len(clusters)})"); tc.font=font(bold=True,size=10,color="FFFFFF"); tc.fill=PatternFill("solid",fgColor=INK)
tc.alignment=Alignment(horizontal="center")
es.merge_cells(start_row=totr,start_column=4,end_row=totr,end_column=6)
es.cell(totr,4,"").fill=PatternFill("solid",fgColor=INK)
for cidx in(2,3,4): es.cell(totr,cidx).border=border

# Ondas / roadmap
wsr=totr+2
sub(es,wsr,"Roadmap de produção (ondas)")
ondas=[("Onda 1 — Fundação (mês 1)","Pilares, comparativos, estudos de dados, CFM-chave, processo. Estabelecem autoridade e capturam quem já busca 'como aparecer no ChatGPT'."),
 ("Onda 2 — Profundidade (mês 2)","Especialidades, condições/sintomas, procedimentos, listicles, casos. Cobrem a intenção do paciente e do médico."),
 ("Onda 3 — Escala local (mês 3)","Páginas programáticas especialidade × cidade. Maior volume, dominam a busca local por IA.")]
hr=wsr+1
for j,h in enumerate(["Onda","Foco"],0):
    c=es.cell(hr,2+j,h); c.font=font(bold=True,color="FFFFFF",size=10); c.fill=PatternFill("solid",fgColor=GOLD)
    c.alignment=Alignment(horizontal="left",vertical="center"); c.border=border
es.merge_cells(start_row=hr,start_column=3,end_row=hr,end_column=6)
for i,(a,b) in enumerate(ondas):
    rr=hr+1+i
    es.cell(rr,2,a.split(" — ")[0]).font=font(size=10,bold=True)
    es.merge_cells(start_row=rr,start_column=3,end_row=rr,end_column=6)
    cc=es.cell(rr,3,a.split("—",1)[1].strip()+": "+b); cc.font=font(size=10); cc.alignment=Alignment(wrap_text=True,vertical="top")
    es.row_dimensions[rr].height=42
    for cidx in (2,3): es.cell(rr,cidx).border=border

# ============================================================
# ABA 3 — ARQUÉTIPOS (Noble + Profound -> médico BR)
# ============================================================
ar=wb.create_sheet("Arquétipos (Noble+Profound)")
ar.sheet_view.showGridLines=False
arch=[
 ("Comparativo 'X vs Y'","Profound vs Ahrefs/Semrush/Conductor (dezenas)","MedRankGPT vs agência tradicional; SEO vs AEO; ChatGPT vs Google","Tabela lado a lado + veredito + FAQ","Decisão; captura quem compara opções"),
 ("Guia prático / playbook","How to rank in ChatGPT; 7-Step Brand Mention Playbook","Como [especialista] aparece no ChatGPT; playbook AEO para consultórios","Passo a passo numerado + checklist + HowTo schema","Educa e converte; muito citável"),
 ("Estudo de dados proprietário","Profound Index; 'how many mentions'","Índice MedRankGPT de especialidades citadas pela IA","Metodologia + gráfico + tabela + Dataset schema","PR + backlinks + citação da IA"),
 ("Roundup de estatísticas","AI Search Statistics 2026: 25 data points","25 estatísticas de IA na saúde","Lista de dados com fonte + FAQ","Ímã de links e citações"),
 ("Mito vs verdade","SEO Isn't Dead. Your Mental Model Is Outdated","SEO morreu para médicos? O que mudou","Mito → realidade + evidência","Engajamento + autoridade"),
 ("Listicle / ferramentas","Top 10 AI Search Monitoring Tools","10 ferramentas de IA para clínicas em 2026","Lista rankeada + ItemList schema","Formato que a IA extrai literalmente"),
 ("Caso de sucesso (% lift)","Avenue Z +250%; 313% lift","Clínica de derma virou citação no ChatGPT","Contexto → ação → resultado em %","Prova social, fundo de funil"),
 ("Glossário / definição","AEO Ultimate Guide; Research Hub","O que é AEO/GEO na medicina","Definição curta e direta + DefinedTerm","Captura buscas 'o que é'; citável"),
 ("Página de processo/transparência","What Happens After You Sign Up for Noble","Como funciona o MedRankGPT","Etapas do serviço + FAQ","Reduz risco; fundo de funil"),
 ("Hook cultural / sazonal","World Cup 2026: Who Wins AI Search","IA na saúde durante campanhas (ex.: Janeiro Branco, Outubro Rosa)","Tema da época + dados + ângulo médico","Tração e novidade"),
 ("Experimento técnico","Markdown vs HTML: AI Traffic experiment","Seu site é legível para a IA? Teste de crawlability médica","Hipótese → teste → resultado","Autoridade técnica; engenharia de conteúdo"),
 ("Programática local","(modelo de escala SEO)","Melhor [especialista] em [cidade] segundo a IA","Template repetível por cidade×especialidade","Volume e domínio da busca local"),
]
ar.merge_cells("B2:F2")
tc=ar.cell(2,2,"Arquétipos de conteúdo — o que Profound e Noble fazem, traduzido para a medicina BR")
tc.font=font(bold=True,size=13,color="FFFFFF"); tc.fill=PatternFill("solid",fgColor=GRAPHITE)
tc.alignment=Alignment(vertical="center",indent=1); ar.row_dimensions[2].height=28
heads=["Arquétipo","Exemplo (Profound/Noble)","Adaptação MedRankGPT","Estrutura / Schema","Por que funciona"]
for j,h in enumerate(heads,2):
    c=ar.cell(3,j,h); c.font=font(bold=True,color="FFFFFF",size=10); c.fill=PatternFill("solid",fgColor=GOLD)
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); c.border=border
ar.row_dimensions[3].height=30
for i,row in enumerate(arch):
    rr=4+i
    for j,v in enumerate(row,2):
        c=ar.cell(rr,j,v); c.font=font(size=9.5); c.alignment=Alignment(wrap_text=True,vertical="top"); c.border=border
    if i%2==0:
        for j in range(2,7): ar.cell(rr,j).fill=PatternFill("solid",fgColor=BAND)
setw(ar,{"A":3,"B":24,"C":30,"D":32,"E":34,"F":34})

# ============================================================
# ABA 4 — REFERÊNCIAS (listas programáticas)
# ============================================================
rf=wb.create_sheet("Referências")
rf.sheet_view.showGridLines=False
rf.merge_cells("B2:E2")
tc=rf.cell(2,2,"Referências para escala programática (especialidades, cidades, padrões de prompt)")
tc.font=font(bold=True,size=13,color="FFFFFF"); tc.fill=PatternFill("solid",fgColor=GRAPHITE)
tc.alignment=Alignment(vertical="center",indent=1); rf.row_dimensions[2].height=28
# especialidades
rf.cell(4,2,"Especialidades-alvo (Tier A)").font=font(bold=True,color="FFFFFF"); rf.cell(4,2).fill=PatternFill("solid",fgColor=GOLD)
for i,e in enumerate(esps): rf.cell(5+i,2,e).font=font(size=10)
# cidades
rf.cell(4,3,"Cidades (busca local)").font=font(bold=True,color="FFFFFF"); rf.cell(4,3).fill=PatternFill("solid",fgColor=GOLD)
for i,c in enumerate(cidades): rf.cell(5+i,3,c).font=font(size=10)
# padrões de prompt
rf.cell(4,4,"Padrões de prompt (a IA recebe)").font=font(bold=True,color="FFFFFF"); rf.cell(4,4).fill=PatternFill("solid",fgColor=GOLD)
prompts=["melhor [especialista] em [cidade]","qual [especialista] tratar [condição]",
 "[procedimento] vale a pena? quem procurar","médico de confiança para [condição] perto de mim",
 "como escolher um bom [especialista]","[especialista] que aceita [convênio] em [cidade]",
 "o que é [condição] e quando procurar médico"]
for i,p in enumerate(prompts):
    rf.merge_cells(start_row=5+i,start_column=4,end_row=5+i,end_column=5)
    rf.cell(5+i,4,p).font=font(size=10); rf.cell(5+i,4).alignment=Alignment(wrap_text=True)
setw(rf,{"A":3,"B":24,"C":20,"D":34,"E":12})
rf.cell(15,2,"Dica: combine [especialidade] × [cidade] e [condição] × [especialidade] para gerar centenas de páginas programáticas a partir destes mesmos modelos.").font=font(italic=True,size=9.5,color="555555")
rf.merge_cells("B15:E15"); rf.row_dimensions[15].height=30
rf.cell(15,2).alignment=Alignment(wrap_text=True,vertical="top")

wb.save("estrategia-100-paginas-medrankgpt.xlsx")
print("OK -> estrategia-100-paginas-medrankgpt.xlsx | linhas:",len(rows))
